"""Compara a planilha do app com a planilha de referência e gera relatório de divergências."""
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ANTIGA = 'Documentos/2026-05/RegimeEspecialMS_2026_05-(2)_1780944108119_tH9y - certa.xlsm'
APP    = 'Documentos/2026-05 fechamento pelo app/RegimeEspecialMS_2026_05_Original.xlsm'
SAIDA  = 'Documentos/2026-05 fechamento pelo app/comparativo_divergencias_2026_05.xlsx'

COLS_COMP = {
    'MOD BC ICMS ST':      'Mod BC',
    'PMC':                 'PMC',
    'MVA':                 'MVA',
    'BC ICMS ST RETIDO':   'BC Retido',
    'VLR ICMS ST RETIDO':  'VLR Retido',
    'BC ICMS ST APURADO':  'BC Apurado',
    'VLR ICMS ST APURADO': 'VLR Apurado',
    'Valor ICMS a Pagar':  'ICMS Pagar',
}


def ler(caminho):
    xl = pd.ExcelFile(caminho)
    df = xl.parse(sheet_name='RegimeEspecial', header=None)
    hdr = [
        str(v).strip() if str(v).strip() not in ('nan', '') else f'COL{i}'
        for i, v in enumerate(df.iloc[3])
    ]
    df = df.iloc[4:].copy()
    df.columns = hdr
    df = df[df['COD PRODUTO NF ORIGEM'].astype(str).str.strip().str.lower() != 'nan']
    df = df.reset_index(drop=True)
    df['_key'] = (
        df['CHAVE DE ACESSO'].astype(str).str.strip()
        + '|'
        + df[hdr[13]].astype(str).str.strip()
    )
    return df


def nv(v):
    if v is None:
        return None
    s = str(v).strip()
    if s.lower() in ('nan', 'none', ''):
        return None
    try:
        return round(float(s.replace(',', '.')), 4)
    except Exception:
        return s


FILL_CRIT   = PatternFill('solid', fgColor='FFB3B3')  # vermelho — ICMS Pagar
FILL_MOD    = PatternFill('solid', fgColor='FFDDB3')  # laranja — MOD diferente
FILL_DIV    = PatternFill('solid', fgColor='FFFF99')  # amarelo — outros campos
FILL_HEADER = PatternFill('solid', fgColor='1F4E79')
FILL_SUBHDR = PatternFill('solid', fgColor='2E75B6')
FONT_W      = Font(color='FFFFFF', bold=True)
BORDA       = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)


def main():
    a = ler(ANTIGA)
    b = ler(APP)
    mapa_a = {row['_key']: row for _, row in a.iterrows()}

    COLS_FIN = list(COLS_COMP.keys())
    COLS_ID  = [
        ('CNPJ',        'CNPJ REMTE'),
        ('Codigo',      'COD PRODUTO NF ORIGEM'),
        ('Descricao',   'DESCRICAO PRODUTO'),
        ('NF',          None),
        ('Item',        None),
        ('Chave NF-e',  'CHAVE DE ACESSO'),
    ]

    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = 'Divergencias'

    # Cabeçalho
    headers = (
        [c[0] for c in COLS_ID]
        + [f'APP\n{v}' for v in COLS_COMP.values()]
        + [f'ANTIGO\n{v}' for v in COLS_COMP.values()]
        + ['Campos divergentes']
    )
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill      = FILL_HEADER
        cell.font      = FONT_W
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 36

    n_diverg = 0
    for _, rb in b.iterrows():
        ra = mapa_a.get(rb['_key'])
        if ra is None:
            continue

        diffs = {}
        for col in COLS_FIN:
            va, vb = nv(ra.get(col)), nv(rb.get(col))
            if va is None and vb is None:
                continue
            try:
                if abs(float(va or 0) - float(vb or 0)) > 0.01:
                    diffs[col] = (va, vb)
            except Exception:
                if str(va) != str(vb):
                    diffs[col] = (va, vb)

        if not diffs:
            continue

        n_diverg += 1
        row_num = n_diverg + 1

        # Identificadores
        id_vals = [
            str(rb.get('CNPJ REMTE', '')),
            str(rb.get('COD PRODUTO NF ORIGEM', '')).lstrip('0') or '0',
            str(rb.get('DESCRICAO PRODUTO', '')),
            str(rb.iloc[4]),
            str(rb.iloc[13]),
            str(rb.get('CHAVE DE ACESSO', '')),
        ]
        for ci, val in enumerate(id_vals, 1):
            ws.cell(row=row_num, column=ci, value=val).border = BORDA

        n_id = len(COLS_ID)
        for fi, col in enumerate(COLS_FIN):
            va, vb = nv(ra.get(col)), nv(rb.get(col))
            c_app = ws.cell(row=row_num, column=n_id + fi + 1, value=vb)
            c_app.border = BORDA
            c_ant = ws.cell(row=row_num, column=n_id + len(COLS_FIN) + fi + 1, value=va)
            c_ant.border = BORDA

            if col in diffs:
                if col == 'Valor ICMS a Pagar':
                    c_app.fill = FILL_CRIT
                    c_ant.fill = FILL_CRIT
                elif col == 'MOD BC ICMS ST':
                    c_app.fill = FILL_MOD
                    c_ant.fill = FILL_MOD
                else:
                    c_app.fill = FILL_DIV
                    c_ant.fill = FILL_DIV

        resumo = '; '.join(
            f'{COLS_COMP[c]}: app={diffs[c][1]} vs ref={diffs[c][0]}'
            for c in diffs
        )
        c_res = ws.cell(row=row_num, column=n_id + 2 * len(COLS_FIN) + 1, value=resumo)
        c_res.border    = BORDA
        c_res.alignment = Alignment(wrap_text=True)

    # Larguras
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 42
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 6
    ws.column_dimensions['F'].width = 46
    for ci in range(7, 7 + 2 * len(COLS_FIN) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.column_dimensions[get_column_letter(7 + 2 * len(COLS_FIN))].width = 80
    ws.freeze_panes = 'A2'

    # ── Aba Resumo ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Resumo')
    ws2['A1'] = 'Resumo das Divergencias - Competencia 2026-05'
    ws2['A1'].font = Font(bold=True, size=13)

    for col, label in zip('ABCDE', ['Tipo de divergencia', 'Qtd itens',
                                     'Soma ICMS Pagar APP (R$)',
                                     'Soma ICMS Pagar Ref (R$)',
                                     'Diferenca (APP - Ref)']):
        c = ws2.cell(row=3, column=ord(col) - 64, value=label)
        c.fill = FILL_SUBHDR
        c.font = FONT_W

    tipos = {
        'MOD BC ICMS ST diferente': [],
        'VLR ICMS ST RETIDO ausente no app': [],
        'BC ICMS ST APURADO diferente': [],
        'Valor ICMS a Pagar divergente (total)': [],
    }

    for _, rb in b.iterrows():
        ra = mapa_a.get(rb['_key'])
        if ra is None:
            continue
        vb_pagar = nv(rb.get('Valor ICMS a Pagar'))
        va_pagar = nv(ra.get('Valor ICMS a Pagar'))

        if nv(ra.get('MOD BC ICMS ST')) != nv(rb.get('MOD BC ICMS ST')):
            tipos['MOD BC ICMS ST diferente'].append((float(vb_pagar or 0), float(va_pagar or 0)))

        ret_a = nv(ra.get('VLR ICMS ST RETIDO'))
        ret_b = nv(rb.get('VLR ICMS ST RETIDO'))
        if ret_a and float(ret_a or 0) > 0 and (ret_b is None or float(ret_b or 0) == 0):
            tipos['VLR ICMS ST RETIDO ausente no app'].append((float(vb_pagar or 0), float(va_pagar or 0)))

        try:
            bc_a = float(nv(ra.get('BC ICMS ST APURADO')) or 0)
            bc_b = float(nv(rb.get('BC ICMS ST APURADO')) or 0)
            if abs(bc_a - bc_b) > 0.01:
                tipos['BC ICMS ST APURADO diferente'].append((float(vb_pagar or 0), float(va_pagar or 0)))
        except Exception:
            pass

        try:
            if abs(float(va_pagar or 0) - float(vb_pagar or 0)) > 0.01:
                tipos['Valor ICMS a Pagar divergente (total)'].append(
                    (float(vb_pagar or 0), float(va_pagar or 0))
                )
        except Exception:
            pass

    for ri, (tipo, itens) in enumerate(tipos.items(), 4):
        s_app = sum(x[0] for x in itens)
        s_ref = sum(x[1] for x in itens)
        ws2.cell(row=ri, column=1, value=tipo)
        ws2.cell(row=ri, column=2, value=len(itens))
        ws2.cell(row=ri, column=3, value=round(s_app, 2))
        ws2.cell(row=ri, column=4, value=round(s_ref, 2))
        c = ws2.cell(row=ri, column=5, value=round(s_app - s_ref, 2))
        if abs(s_app - s_ref) > 1:
            c.fill = FILL_CRIT if s_app > s_ref else FILL_DIV

    ws2.column_dimensions['A'].width = 44
    ws2.column_dimensions['B'].width = 12
    for col in 'CDE':
        ws2.column_dimensions[col].width = 24

    # Nota explicativa
    ws2['A9']  = 'Legenda de cores (aba Divergencias):'
    ws2['A9'].font = Font(bold=True)
    for row, (fill, texto) in enumerate([
        (FILL_CRIT, 'Vermelho — Valor ICMS a Pagar divergente'),
        (FILL_MOD,  'Laranja  — MOD BC ICMS ST diferente'),
        (FILL_DIV,  'Amarelo  — Outros campos (PMC, MVA, BC, VLR)'),
    ], 10):
        ws2.cell(row=row, column=1).fill  = fill
        ws2.cell(row=row, column=2, value=texto)

    wb.save(SAIDA)
    print(f'Salvo: {SAIDA}')
    print(f'Linhas com divergencia: {n_diverg}')
    print()
    for tipo, itens in tipos.items():
        s_app = sum(x[0] for x in itens)
        s_ref = sum(x[1] for x in itens)
        print(f'  {tipo}: {len(itens)} itens | APP={s_app:.2f} | Ref={s_ref:.2f} | Dif={s_app-s_ref:.2f}')


if __name__ == '__main__':
    main()
