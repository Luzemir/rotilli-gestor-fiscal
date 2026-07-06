"""
preparar_template_regime_especial.py — Script pontual de preparação do template

Gera templates/RegimeEspecialMS_template.xlsm a partir de um arquivo histórico
real, esvaziando as linhas de dados das abas RegimeEspecial e RelaçãoNFE mas
preservando GNRE vazia, Planilha1 e as tabelas estáticas de referência
(Supérfluos, ListaMVA, MVA).

Revisão 2026-06-16: removida a coluna "Coluna2" (AC, sem uso), deslocando as
colunas de cálculo uma posição à esquerda. Adicionadas duas colunas novas:
"PMC CMED" (AN) e "Diferença PMC" (AO), com "Coluna3" (AM) como preenchimento
para essas duas caírem exatamente nas letras AN/AO pedidas. A coluna "PMC" (Y)
passa a conter sempre o PMC histórico do produto (nunca o da CMED).

Rodar de novo só se for necessário atualizar as tabelas de referência
(Supérfluos/ListaMVA) quando a legislação mudar, ou trocar o arquivo-base.

Uso:
  python scripts/preparar_template_regime_especial.py
"""
import shutil
import openpyxl
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

ORIGEM = 'Documentos/2026-04/RegimeEspecialMS_2026_04-Original.xlsm'
DESTINO = 'templates/RegimeEspecialMS_template.xlsm'

# Cabeçalho final da aba RegimeEspecial (41 colunas, A a AO)
NOVOS_CABECALHOS = [
    'ANO', 'MÊS', 'DATA', 'CHAVE DE ACESSO', 'Nº NF', 'RAZÃO REMTE', 'CNPJ REMTE',
    'UF REMTE', 'CEST', 'COD NCM', 'COD EAN GTIN TRIB', 'COD CFOP',
    'COD PRODUTO NF ORIGEM', 'Nº ITEM', 'DESCRICAO PRODUTO', 'UNID.', 'QTDE',
    'VLR UNIT TRIB.', 'FRETE', 'SEGURO', 'IPI/OUTRAS DESPESAS ACESSÓRIAS',
    'DESCONTO', 'BC ICMS (ORIGEM)', 'VLR ICMS (ORIGEM)', 'PMC', 'MVA',
    'MOD BC ICMS ST', 'BC ICMS ST RETIDO',
    'VLR ICMS ST RETIDO', 'BC ICMS ST APURADO', 'Aliquota Interna',
    'VLR ICMS ST APURADO', 'CRÉDITO OUTORGADO É DEVIDO? (S ou N)', 'CALC. 1',
    'CALC. 2', 'CRÉD. OUTORG.', 'Valor ICMS a Pagar',
    'Coluna1', 'Coluna3', 'PMC CMED', 'Diferença PMC',
]
assert len(NOVOS_CABECALHOS) == 41


def main():
    shutil.copy2(ORIGEM, DESTINO)
    wb = openpyxl.load_workbook(DESTINO, keep_vba=True, data_only=False)

    # ── RegimeEspecial ────────────────────────────────────────────────────
    ws = wb['RegimeEspecial']
    estilo_antigo = ws.tables['tabRegEsp'].tableStyleInfo
    del ws.tables['tabRegEsp']

    # Limpa todas as células (cabeçalho + dados) das 41 colunas
    for row in range(4, 515):
        for col in range(1, 42):
            ws.cell(row=row, column=col).value = None

    # Escreve o novo cabeçalho na linha 4
    for col, nome in enumerate(NOVOS_CABECALHOS, 1):
        ws.cell(row=4, column=col).value = nome

    tabela_nova = Table(displayName='tabRegEsp', ref='A4:AO5')
    tabela_nova.tableColumns = [
        TableColumn(id=i, name=nome) for i, nome in enumerate(NOVOS_CABECALHOS, 1)
    ]
    tabela_nova.tableStyleInfo = estilo_antigo or TableStyleInfo(
        name='TableStyleMedium23', showRowStripes=True, showColumnStripes=True
    )
    ws.add_table(tabela_nova)

    # ── RelaçãoNFE: limpa linhas de dados (2 a 62), mantém cabeçalho linha 1 ─
    ws2 = wb['RelaçãoNFE']
    tbl2 = ws2.tables['tabRelNFE']
    for row in range(2, 63):
        for col in range(1, 12):  # A..K
            ws2.cell(row=row, column=col).value = None
    tbl2.ref = 'A1:I2'
    if tbl2.autoFilter is not None:
        tbl2.autoFilter.ref = 'A1:I2'

    # ── Resumo: corrige referências às colunas que se deslocaram ────────────
    ws_resumo = wb['Resumo']
    ws_resumo['C10'].value = '=SUM(RegimeEspecial!AD:AD)'  # BC ICMS ST APURADO (era AE)
    ws_resumo['C11'].value = '=SUM(RegimeEspecial!AF:AF)'  # VLR ICMS ST APURADO (era AG)
    ws_resumo['C13'].value = '=SUM(RegimeEspecial!AJ:AJ)'  # CRÉD. OUTORG. (era AK)
    # C14 usa referência estruturada tabRegEsp[VLR ICMS ST RETIDO] — robusta ao deslocamento, não precisa mudar

    wb.save(DESTINO)
    print(f'Template salvo em {DESTINO}')


if __name__ == '__main__':
    main()
