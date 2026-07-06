import math
import io
import pandas as pd
import openpyxl
import streamlit as st

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from _config import aplicar_tema_contili, page_header

# ── Constantes ────────────────────────────────────────────────────────────────

HEADER_ROW = 4
DATA_START  = 5   # app gera dados a partir da linha 5

C_CHAVE = 4   # D  — CHAVE DE ACESSO
C_ITEM  = 14  # N  — Nº ITEM
C_NF    = 5   # E  — Nº NF
C_PROD  = 13  # M  — COD PRODUTO NF ORIGEM
C_DESC  = 15  # O  — DESCRICAO PRODUTO
C_AK    = 37  # AK — Valor ICMS a Pagar

# Grupos usados para descrever o motivo da divergência
EXPLAIN_GROUPS = [
    ("Dados básicos da NF", [
        (5,  "Nº NF"),
        (15, "Descrição"),
        (17, "Qtde"),
        (18, "Vlr Unit"),
    ]),
    ("Entradas de cálculo", [
        (19, "Frete"),
        (20, "Seguro"),
        (21, "IPI/Outras Despesas"),
        (22, "Desconto"),
        (23, "BC ICMS Origem"),
        (24, "Vlr ICMS Origem"),
    ]),
    ("Parâmetros ST", [
        (25, "PMC"),
        (26, "MVA"),
        (27, "Mod BC ST"),
    ]),
    ("ST Retido (NF origem)", [
        (28, "BC ST Retido"),
        (29, "Vlr ST Retido"),
    ]),
    ("Crédito Outorgado", [
        (33, "Créd Devido?"),
        (34, "Cálc 1"),
        (35, "Cálc 2"),
        (36, "Créd Outorgado"),
    ]),
]

# ── Funções de lógica ─────────────────────────────────────────────────────────

def norm(v):
    """Normaliza para comparação: None / nan / 0 → None."""
    if v is None:
        return None
    if isinstance(v, float):
        if math.isnan(v):
            return None
        r = round(v, 4)
        return None if r == 0.0 else r
    if isinstance(v, int):
        return None if v == 0 else v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            f = float(s)
            if f == 0:
                return None
            return int(f) if f == int(f) else round(f, 4)
        except ValueError:
            return s
    return v


def fmt(v):
    if v is None:
        return "—"
    if isinstance(v, float):
        return f"{v:,.4f}".rstrip("0").rstrip(".")
    return str(v)


def resolve_ak(row):
    """
    Retorna o valor de AK (Valor ICMS a Pagar).
    Usa o valor em cache quando disponível; caso contrário calcula
    manualmente: AK = AF − AJ − AC.
    """
    ak = row[C_AK - 1]
    if ak is not None and not (isinstance(ak, str) and ak.startswith("=")):
        try:
            return float(ak)
        except (TypeError, ValueError):
            pass
    af = row[31]   # col AF (32 → índice 31)
    aj = row[35]   # col AJ (36 → índice 35)
    ac = row[28]   # col AC (29 → índice 28)
    try:
        return (float(af) or 0) - (float(aj) or 0) - (float(ac) or 0)
    except (TypeError, ValueError):
        return None


def load_sheet(uploaded_file) -> dict:
    """
    Lê a aba RegimeEspecial e retorna um dict:
        (chave_acesso, num_item) → tuple com todos os valores da linha.
    Aceita UploadedFile do Streamlit ou caminho de arquivo.
    Linhas sem CHAVE DE ACESSO são ignoradas (linhas de totais / vazias).
    """
    if hasattr(uploaded_file, 'read'):
        uploaded_file.seek(0)
        source = io.BytesIO(uploaded_file.read())
    else:
        source = uploaded_file

    wb = openpyxl.load_workbook(source, keep_vba=True, read_only=True, data_only=True)
    ws = wb["RegimeEspecial"]
    index = {}
    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        chave = row[C_CHAVE - 1]
        item  = row[C_ITEM  - 1]
        if chave is None:
            continue
        key = (str(chave).strip(), item)
        index[key] = row
    wb.close()
    return index


def _build_motivo(orig_row, ajus_row) -> str:
    parts = []
    for grupo, cols in EXPLAIN_GROUPS:
        diffs = []
        for ci, label in cols:
            ov = norm(orig_row[ci - 1] if ci - 1 < len(orig_row) else None)
            av = norm(ajus_row[ci - 1] if ci - 1 < len(ajus_row) else None)
            if ov != av:
                diffs.append(f"{label}: App={fmt(ov)} / Ajustada={fmt(av)}")
        if diffs:
            parts.append(f"[{grupo}] " + " | ".join(diffs))
    if not parts:
        parts.append("Valor ICMS a Pagar difere — causa não identificada nos campos mapeados.")
    return "\n".join(parts)


def comparar(app_index: dict, ajus_index: dict) -> list:
    rows = []
    all_keys = sorted(set(app_index) | set(ajus_index))

    for key in all_keys:
        app_row  = app_index.get(key)
        ajus_row = ajus_index.get(key)
        chave, item = key

        if ajus_row is None:
            rows.append({
                "Chave de Acesso":       chave,
                "NF":                    app_row[C_NF - 1] if app_row else None,
                "Item":                  item,
                "Cód Produto":           app_row[C_PROD - 1] if app_row else None,
                "Descrição":             app_row[C_DESC - 1] if app_row else None,
                "ICMS Pagar – App":      resolve_ak(app_row) if app_row else None,
                "ICMS Pagar – Ajustada": None,
                "Diferença":             None,
                "Status":                "⚠️ Ausente na Ajustada",
                "Motivo":                "Linha presente no App mas ausente na planilha Ajustada.",
            })
            continue

        if app_row is None:
            rows.append({
                "Chave de Acesso":       chave,
                "NF":                    ajus_row[C_NF - 1],
                "Item":                  item,
                "Cód Produto":           ajus_row[C_PROD - 1],
                "Descrição":             ajus_row[C_DESC - 1],
                "ICMS Pagar – App":      None,
                "ICMS Pagar – Ajustada": resolve_ak(ajus_row),
                "Diferença":             None,
                "Status":                "⚠️ Linha extra na Ajustada",
                "Motivo":                "Linha presente na Ajustada mas ausente no App.",
            })
            continue

        # Validação cruzada: mesmo produto na mesma NF?
        prod_app  = str(app_row[C_PROD - 1]).strip() if app_row[C_PROD - 1] is not None else ""
        prod_ajus = str(ajus_row[C_PROD - 1]).strip() if ajus_row[C_PROD - 1] is not None else ""
        nf_app    = str(app_row[C_NF - 1]).strip() if app_row[C_NF - 1] is not None else ""
        nf_ajus   = str(ajus_row[C_NF - 1]).strip() if ajus_row[C_NF - 1] is not None else ""
        avisos_key = []
        if prod_app != prod_ajus:
            avisos_key.append(f"⚠️ COD PRODUTO DIVERGE (App={prod_app} / Ajustada={prod_ajus})")
        if nf_app != nf_ajus:
            avisos_key.append(f"⚠️ Nº NF DIVERGE (App={nf_app} / Ajustada={nf_ajus})")
        prefixo_aviso = ("\n".join(avisos_key) + "\n") if avisos_key else ""

        ak_app  = resolve_ak(app_row)
        ak_ajus = resolve_ak(ajus_row)

        if norm(ak_app) == norm(ak_ajus):
            rows.append({
                "Chave de Acesso":       chave,
                "NF":                    ajus_row[C_NF - 1],
                "Item":                  item,
                "Cód Produto":           ajus_row[C_PROD - 1],
                "Descrição":             ajus_row[C_DESC - 1],
                "ICMS Pagar – App":      ak_app,
                "ICMS Pagar – Ajustada": ak_ajus,
                "Diferença":             0.0,
                "Status":                "✅ OK",
                "Motivo":                prefixo_aviso,
            })
        else:
            motivo = prefixo_aviso + _build_motivo(app_row, ajus_row)
            diff   = (ak_ajus or 0) - (ak_app or 0)
            rows.append({
                "Chave de Acesso":       chave,
                "NF":                    ajus_row[C_NF - 1],
                "Item":                  item,
                "Cód Produto":           ajus_row[C_PROD - 1],
                "Descrição":             ajus_row[C_DESC - 1],
                "ICMS Pagar – App":      ak_app,
                "ICMS Pagar – Ajustada": ak_ajus,
                "Diferença":             round(diff, 4),
                "Status":                "🔴 Divergente",
                "Motivo":                motivo,
            })

    return rows


# ── Página ────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Comparar Planilhas", page_icon="📊", layout="wide")
aplicar_tema_contili()

page_header(
    'Comparar Planilhas',
    'Confronta a planilha Original com a Ajustada e identifica linha a linha o que mudou no ICMS a Pagar.',
)

# ── Upload ────────────────────────────────────────────────────────────────────

col1, col2 = st.columns(2)
with col1:
    st.markdown("**📂 Planilha do App** *(base — gerada antes do ajuste)*")
    f_app = st.file_uploader("App", type=["xlsm", "xlsx"], label_visibility="collapsed")
with col2:
    st.markdown("**📂 Planilha Ajustada** *(fiscal — retornada pelo contador)*")
    f_ajus = st.file_uploader("Ajustada", type=["xlsm", "xlsx"], label_visibility="collapsed")

if not f_app or not f_ajus:
    st.info("Faça o upload das duas planilhas para iniciar a comparação.")
    st.stop()

# ── Comparação ────────────────────────────────────────────────────────────────

with st.spinner("Carregando e comparando planilhas..."):
    try:
        app_idx  = load_sheet(f_app)
        ajus_idx = load_sheet(f_ajus)
    except KeyError:
        st.error("Uma das planilhas não contém a aba **RegimeEspecial**. Verifique os arquivos.")
        st.stop()
    except Exception as e:
        st.error(f"Erro ao abrir planilha: {e}")
        st.stop()

    resultado = comparar(app_idx, ajus_idx)

df = pd.DataFrame(resultado)

# ── Métricas ──────────────────────────────────────────────────────────────────

total      = len(df)
ok         = int((df["Status"] == "✅ OK").sum())
divergente = int((df["Status"] == "🔴 Divergente").sum())
ausente    = int(df["Status"].str.contains("Ausente|extra").sum())
soma_diff  = df["Diferença"].sum()

m1, m2, m3, m4, m5 = st.columns(5)
m1.metric("Total linhas",        total)
m2.metric("✅ Iguais",           ok)
m3.metric("🔴 Divergentes",     divergente)
m4.metric("⚠️ Ausentes / extras", ausente)
m5.metric("Σ Diferença ICMS",   f"R$ {soma_diff:,.2f}")

st.divider()

# ── Filtro ────────────────────────────────────────────────────────────────────

filtro = st.radio(
    "Mostrar:",
    ["Todas", "Apenas divergentes", "Apenas OK"],
    horizontal=True,
)
if filtro == "Apenas divergentes":
    df_view = df[df["Status"] != "✅ OK"]
elif filtro == "Apenas OK":
    df_view = df[df["Status"] == "✅ OK"]
else:
    df_view = df

st.caption(f"{len(df_view)} linha(s) exibida(s)")

# ── Tabela ────────────────────────────────────────────────────────────────────

def _highlight(row):
    if row["Status"] == "🔴 Divergente":
        return ["background-color: #ffe0e0"] * len(row)
    if "Ausente" in str(row["Status"]) or "extra" in str(row["Status"]):
        return ["background-color: #fff3cd"] * len(row)
    return [""] * len(row)

st.dataframe(
    df_view.style.apply(_highlight, axis=1),
    use_container_width=True,
    height=520,
    column_config={
        "Chave de Acesso":       st.column_config.TextColumn("Chave de Acesso", width="medium"),
        "ICMS Pagar – App":      st.column_config.NumberColumn("ICMS Pagar – App",      format="R$ %.4f"),
        "ICMS Pagar – Ajustada": st.column_config.NumberColumn("ICMS Pagar – Ajustada", format="R$ %.4f"),
        "Diferença":             st.column_config.NumberColumn("Diferença",              format="R$ %.4f"),
        "Motivo":                st.column_config.TextColumn("Motivo", width="large"),
        "Status":                st.column_config.TextColumn("Status", width="small"),
    },
)

# ── Download ──────────────────────────────────────────────────────────────────

st.divider()
_NUM_COLS = {"ICMS Pagar – App", "ICMS Pagar – Ajustada", "Diferença"}
buf = io.BytesIO()
with pd.ExcelWriter(buf, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Comparativo")
    ws_xls = writer.sheets["Comparativo"]
    ws_xls.freeze_panes = "A2"
    col_names = list(df.columns)
    for col_idx, col_name in enumerate(col_names, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if col_name in _NUM_COLS:
            for row_num in range(2, len(df) + 2):
                ws_xls[f"{col_letter}{row_num}"].number_format = "#,##0.00"
        if col_name == "Chave de Acesso":
            ws_xls.column_dimensions[col_letter].width = 46
        elif col_name == "Motivo":
            ws_xls.column_dimensions[col_letter].width = 80
        elif col_name == "Descrição":
            ws_xls.column_dimensions[col_letter].width = 30

st.download_button(
    label="⬇️ Baixar relatório completo (.xlsx)",
    data=buf.getvalue(),
    file_name="comparativo_regime_especial.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
