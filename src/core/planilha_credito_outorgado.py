"""
planilha_credito_outorgado.py — Geração da planilha de apuração do ICMS-ST
por Crédito Outorgado (trilha Original)

Lê todos os itens de NF-e já persistidos em nfe_item_apuracao para uma
competência, clona o template (templates/RegimeEspecialMS_template.xlsm) e
preenche RegimeEspecial (dados literais + fórmulas vivas, replicadas
literalmente da planilha histórica) e RelaçãoNFE (uma linha por NF-e única).
A aba Resumo já é formula-driven (soma colunas inteiras) — só o Período é
atualizado por geração.

Reentrante/idempotente: cada chamada relê a competência inteira do banco e
clona o template do zero, então gerar de novo no mesmo mês apenas sobrescreve
o arquivo anterior com o snapshot atual.
"""

import os
import datetime
import openpyxl

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
TEMPLATE_PATH = os.path.join(ROOT, 'templates', 'RegimeEspecialMS_template.xlsm')

LINHA_INICIAL_DADOS = 5  # RegimeEspecial: cabeçalho na linha 4, dados a partir da 5
LINHA_INICIAL_RELNFE = 2  # RelaçãoNFE: cabeçalho na linha 1, dados a partir da 2

# Colunas A-AA + W,X: dados literais. Mapeamento coluna -> chave do registro do banco
# (None = derivado separadamente, ver _escrever_linha_regime_especial)
_COLUNAS_LITERAIS = {
    'A': None,  # ANO (derivado de data_emissao)
    'B': None,  # MÊS (derivado de data_emissao)
    'C': None,  # DATA (derivado de data_emissao)
    'D': 'chave_acesso',
    'E': 'num_nf',
    'F': 'razao_remetente',
    'G': 'cnpj_remetente',
    'H': 'uf_remetente',
    'I': 'cest',
    'J': 'ncm',
    'K': 'ean',
    'L': 'cfop',
    'M': 'cod_produto_origem',
    'N': 'num_item',
    'O': 'descricao_produto',
    'P': 'unidade',
    'Q': 'quantidade',
    'R': 'valor_unitario',
    'S': 'frete',
    'T': 'seguro',        # Seguro + Despesas Acessórias (vOutro) somados — col "SEGURO/DESP. ACESS."
    'U': 'ipi_despesas',  # IPI do produto — col "IPI"
    'V': 'desconto',
    'W': 'bc_icms_origem',
    'X': 'vlr_icms_origem',
    'Y': 'pmc',         # PMC histórico do produto — NUNCA o da CMED (esse vai em AN)
    'Z': 'mva',
    'AA': None,  # MOD BC ICMS ST (convertido para int)
    'AB': 'bc_icms_st_nfe',   # BC ICMS ST declarado pelo remetente na NF-e
    'AC': 'vlr_icms_st_nfe',  # VLR ICMS ST declarado pelo remetente na NF-e
    'AN': 'pmc_cmed',   # PMC pesquisado na CMED (não confundir com Y)
}

_FORMATOS_NUMERO = {
    'C': 'mm-dd-yy',
    'D': '@', 'G': '@', 'I': '@', 'J': '@', 'K': '@',
    'R': '#,##0.00_);[Red]\\(\\-\\)', 'S': '#,##0.00_);[Red]\\(\\-\\)',
    'T': '#,##0.00_);[Red]\\(\\-\\)', 'U': '#,##0.00_);[Red]\\(\\-\\)',
    'V': '#,##0.00_);[Red]\\(\\-\\)', 'W': '#,##0.00_);[Red]\\(\\-\\)',
    'X': '#,##0.00_);[Red]\\(\\-\\)',
    'Y': '_-* #,##0.00_-;_-* "-"??_-',
    'Z': '_-* #,##0.0000_-',
    'AB': '#,##0.00_);[Red]\\(\\-\\)', 'AC': '#,##0.00_);[Red]\\(\\-\\)',
    'AD': '#,##0.00_);[Red]\\(\\-\\)', 'AF': '#,##0.00_);[Red]\\(\\-\\)',
    'AH': '#,##0.00_);[Red]\\(\\-\\)', 'AI': '#,##0.00_);[Red]\\(\\-\\)',
    'AJ': '#,##0.00_);[Red]\\(\\-\\)', 'AK': '#,##0.00_);[Red]\\(\\-\\)',
    'AN': '_-* #,##0.00_-;_-* "-"??_-',
    'AO': '#,##0.00_);[Red]\\(\\-\\)',
    'AP': '0.00%',
}

# Fórmulas literais (replicadas exatamente da planilha histórica, com as
# referências internas já ajustadas ao novo layout sem "Coluna2"), {n} = linha
_FORMULAS = {
    'AD': (
        '=IF(AA{n}=0,IF(AND(Y{n}>0,Z{n}=0),Q{n}*Y{n},"PMC ou MVA ?"),'
        'IF(AA{n}=4,IF(AND(Y{n}=0,Z{n}>0,Z{n}<=5),'
        '((Q{n}*R{n}+S{n}+T{n}+U{n}-V{n})*(1+Z{n})),"PMC or MVA<5?"),'
        'IF(AA{n}=5,IF(OR(Y{n}>0,Z{n}>0),"#Limpar PMC/MVA-TribNormal#","TribNormal- 0,00"),'
        'IF(OR(AA{n}=1,AA{n}=2,AA{n}=3),IF(AND(Y{n}=0,Z{n}=0),'
        '((Q{n}*R{n}+S{n}+T{n}+U{n}-V{n})*VLOOKUP('
        'IF(X{n}/W{n}<0.058,0.04,IF(X{n}/W{n}<0.09,0.07,0.12))+AA{n},'
        'ListaMVA[],5,FALSE)),"#Limpar PMC/MVA-Listas"),"err!"))))'
    ),
    'AE': '=IF(ISNUMBER(MATCH(J{n},INDEX(Superfluos[],,0),0)),20%,17%)',
    'AF': (
        '=IF(AD{n}="TribNormal- 0,00",0,IF(AG{n}="S",'
        'IF(X{n}/W{n}<0.045,(AD{n}*AE{n})-(Q{n}*R{n}+S{n}+T{n}+U{n}-V{n})*0.04,'
        '(AD{n}*AE{n})-(Q{n}*R{n}+S{n}+T{n}+U{n}-V{n})*0.07),(AD{n}*AE{n})-X{n}))'
    ),
    'AG': '=IF(AND(H{n}<>"MS",AA{n}=0),"S","N")',
    'AH': (
        '=IF(AG{n}="S",(((Q{n}*R{n})+S{n}+T{n}+U{n}-V{n})*1.983*0.153)-'
        '(((Q{n}*R{n})+S{n}+T{n}+U{n}-V{n})*0.07),0)'
    ),
    'AI': '=IF(AG{n}="S",(Q{n}*Y{n}*0.6*0.153)-(((Q{n}*R{n})+S{n}+T{n}+U{n}-V{n})*0.07),0)',
    'AJ': (
        '=IF(AG{n}="S",IF(OR(AH{n}<=0,AI{n}<=0),#VALUE!,'
        'IF(MAX(AH{n},AI{n})>AF{n},AF{n},AF{n}-MAX(AH{n},AI{n}))),0)'
    ),
    'AK': '=AF{n}-AJ{n}-AC{n}',
    'AO': '=IF(AN{n}<>Y{n},AN{n}-Y{n},0)',
    # Alíquota de ICMS efetivamente aplicada na origem do produto (VLR ICMS / BC ICMS)
    'AP': '=IF(W{n}=0,0,X{n}/W{n})',
}


def _col_idx(letra):
    return openpyxl.utils.column_index_from_string(letra)


def _normalizar_mva(mva):
    """
    Garante que o MVA esteja sempre no formato fracionário (ex: 0,3824 para 38,24%).
    Se alguém digitou o percentual cheio por engano (ex: 38.24 ou 60), o valor
    fica > 5 (500% em fração) — acima do que a fórmula da coluna AD aceita
    (Z<=5) — então é corrigido dividindo por 100.
    """
    if mva is None:
        return None
    mva = float(mva)
    if mva > 5:
        mva = mva / 100
    return round(mva, 4)


def _data_emissao_partes(data_emissao):
    """Extrai (ano, mes, date) de uma string ISO de data_emissao da NF-e."""
    if not data_emissao:
        return None, None, None
    texto = data_emissao[:10]  # 'YYYY-MM-DD'
    try:
        dt = datetime.datetime.strptime(texto, '%Y-%m-%d')
        return dt.year, dt.month, dt.date()
    except ValueError:
        return None, None, None


def _buscar_itens(conn, mes_ano, tipo_planilha):
    cursor = conn.cursor()
    cursor.execute(
        '''SELECT chave_acesso, num_item, num_nf, data_emissao, cnpj_remetente,
                  razao_remetente, uf_remetente, cest, ncm, ean, cfop,
                  cod_produto_origem, descricao_produto, unidade, quantidade,
                  valor_unitario, frete, seguro, ipi_despesas, desconto,
                  bc_icms_origem, vlr_icms_origem, pmc, pmc_cmed, mva, mod_bc_icms_st,
                  bc_icms_st_nfe, vlr_icms_st_nfe
           FROM nfe_item_apuracao
           WHERE mes_ano = ? AND tipo_planilha = ?
           ORDER BY data_emissao, num_nf, num_item''',
        (mes_ano, tipo_planilha),
    )
    colunas = [d[0] for d in cursor.description]
    return [dict(zip(colunas, row)) for row in cursor.fetchall()]


def _escrever_linha_regime_especial(ws, linha, item):
    ano, mes, data = _data_emissao_partes(item.get('data_emissao'))
    ws.cell(row=linha, column=_col_idx('A')).value = ano
    ws.cell(row=linha, column=_col_idx('B')).value = mes
    ws.cell(row=linha, column=_col_idx('C')).value = data

    for letra, campo in _COLUNAS_LITERAIS.items():
        if campo is None:
            continue
        valor = item.get(campo)
        if letra == 'Z':
            valor = _normalizar_mva(valor)
        ws.cell(row=linha, column=_col_idx(letra)).value = valor

    mod_bc = item.get('mod_bc_icms_st')
    ws.cell(row=linha, column=_col_idx('AA')).value = int(mod_bc) if mod_bc is not None else None

    for letra, fmt in _FORMATOS_NUMERO.items():
        ws.cell(row=linha, column=_col_idx(letra)).number_format = fmt

    for letra, template in _FORMULAS.items():
        ws.cell(row=linha, column=_col_idx(letra)).value = template.format(n=linha)


def _agrupar_por_nf(itens):
    """Agrupa itens por chave_acesso, somando valores para a aba RelaçãoNFE."""
    grupos = {}
    for item in itens:
        chave = item.get('chave_acesso')
        if chave not in grupos:
            grupos[chave] = {
                'chave_acesso': chave,
                'num_nf': item.get('num_nf'),
                'razao_remetente': item.get('razao_remetente'),
                'data_emissao': item.get('data_emissao'),
                'valor_total': 0.0,
                'bc_icms_origem': 0.0,
                'vlr_icms_origem': 0.0,
            }
        g = grupos[chave]
        valor_unit = item.get('valor_unitario') or 0.0
        qtde = item.get('quantidade') or 0.0
        g['valor_total'] += valor_unit * qtde
        g['bc_icms_origem'] += item.get('bc_icms_origem') or 0.0
        g['vlr_icms_origem'] += item.get('vlr_icms_origem') or 0.0
    return list(grupos.values())


def gerar_planilha(conn, mes_ano, tipo_planilha='Original'):
    """
    Gera a planilha de apuração para a competência informada, lendo todos os
    itens já persistidos em nfe_item_apuracao (acumulado de todas as sessões
    de importação do mês). Retorna o caminho absoluto do arquivo gerado.
    """
    itens = _buscar_itens(conn, mes_ano, tipo_planilha)

    wb = openpyxl.load_workbook(TEMPLATE_PATH, keep_vba=True, data_only=False)

    # --- RegimeEspecial ---
    ws = wb['RegimeEspecial']
    for idx, item in enumerate(itens):
        linha = LINHA_INICIAL_DADOS + idx
        _escrever_linha_regime_especial(ws, linha, item)

    ultima_linha = LINHA_INICIAL_DADOS + max(len(itens), 1) - 1
    novo_ref = f'A4:AP{ultima_linha}'
    ws.tables['tabRegEsp'].ref = novo_ref
    if ws.tables['tabRegEsp'].autoFilter is not None:
        ws.tables['tabRegEsp'].autoFilter.ref = novo_ref

    # --- RelaçãoNFE ---
    ws2 = wb['RelaçãoNFE']
    grupos = _agrupar_por_nf(itens)
    for idx, g in enumerate(grupos):
        linha = LINHA_INICIAL_RELNFE + idx
        _, _, data = _data_emissao_partes(g.get('data_emissao'))
        ws2.cell(row=linha, column=1).value = data
        ws2.cell(row=linha, column=2).value = g['chave_acesso']
        ws2.cell(row=linha, column=3).value = g['num_nf']
        ws2.cell(row=linha, column=4).value = g['razao_remetente']
        ws2.cell(row=linha, column=5).value = round(g['valor_total'], 2)
        ws2.cell(row=linha, column=6).value = round(g['bc_icms_origem'], 2)
        ws2.cell(row=linha, column=7).value = round(g['vlr_icms_origem'], 2)
        # H (ICMS ST Retido) fica em branco — sem fonte de dados própria ainda
        ws2.cell(row=linha, column=9).value = (
            '=SUMIF(RegimeEspecial!D:D,tabRelNFE[[#This Row],[CHAVE DE ACESSO]],RegimeEspecial!AK:AK)'
        )

    ultima_linha_nfe = LINHA_INICIAL_RELNFE + max(len(grupos), 1) - 1
    novo_ref_nfe = f'A1:I{ultima_linha_nfe}'
    ws2.tables['tabRelNFE'].ref = novo_ref_nfe
    if ws2.tables['tabRelNFE'].autoFilter is not None:
        ws2.tables['tabRelNFE'].autoFilter.ref = novo_ref_nfe

    # --- Resumo ---
    ano, mes = int(mes_ano[:4]), int(mes_ano[5:7])
    ws_res = wb['Resumo']
    ws_res['B6'].value = datetime.datetime(ano, mes, 1)
    # Fórmulas com referência estruturada de tabela (Excel 2007+):
    # somam apenas o corpo de dados do tabRegEsp, evitando duplicidade
    # com eventuais linhas de totais fora da tabela.
    ws_res['C10'] = '=SUM(tabRegEsp[BC ICMS ST APURADO])'
    ws_res['C11'] = '=SUM(tabRegEsp[VLR ICMS ST APURADO])'
    ws_res['C13'] = '=SUM(tabRegEsp[CRÉD. OUTORG.])'
    ws_res['C14'] = '=SUM(GNRE[VALOR GNRE])+SUM(tabRegEsp[VLR ICMS ST RETIDO])'

    # --- Salvar ---
    pasta_destino = os.path.join(ROOT, 'Documentos', mes_ano)
    os.makedirs(pasta_destino, exist_ok=True)
    nome_arquivo = f'RegimeEspecialMS_{ano}_{mes:02d}_{tipo_planilha}.xlsm'
    caminho = os.path.join(pasta_destino, nome_arquivo)
    wb.save(caminho)

    return caminho
