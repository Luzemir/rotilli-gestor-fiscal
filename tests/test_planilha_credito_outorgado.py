"""
Testes da geração da planilha de apuração (planilha_credito_outorgado.py).

Cobre:
  - _escrever_linha_regime_especial: colunas literais (T=Seguro/Desp.Acess.,
    U=IPI), normalização do MVA na coluna Z, MOD como int em AA, fórmulas
    vivas (AD, AP) e formatos numéricos.
  - gerar_planilha ponta a ponta com banco em memória e saída em pasta
    temporária: range da tabela tabRegEsp, consistência célula×metadata
    dos cabeçalhos (armadilha do Excel Table — ver AGENTS.md).

Execução:
    python -m pytest tests/test_planilha_credito_outorgado.py -v
"""
import sys
import os
import sqlite3
import pytest
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src', 'core'))

import planilha_credito_outorgado as p


def _item_base(**sobrescreve):
    item = {
        'chave_acesso': '31260400000000000000550010000000011000000001',
        'num_item': 1, 'num_nf': '123', 'data_emissao': '2026-05-15T10:00:00-03:00',
        'cnpj_remetente': '00000000000191', 'razao_remetente': 'FORNECEDOR TESTE',
        'uf_remetente': 'MG', 'cest': '13.001.00', 'ncm': '30049099',
        'ean': '7891000315507', 'cfop': '6102', 'cod_produto_origem': 'P001',
        'descricao_produto': 'PRODUTO TESTE', 'unidade': 'UN',
        'quantidade': 10.0, 'valor_unitario': 5.0,
        'frete': 1.0, 'seguro': 2.5, 'ipi_despesas': 3.3, 'desconto': 0.5,
        'bc_icms_origem': 100.0, 'vlr_icms_origem': 7.0,
        'pmc': 0.0, 'pmc_cmed': 0.0, 'mva': 0.3824, 'mod_bc_icms_st': '4',
        'bc_icms_st_nfe': 0.0, 'vlr_icms_st_nfe': 0.0,
    }
    item.update(sobrescreve)
    return item


@pytest.fixture(scope='module')
def template():
    return openpyxl.load_workbook(p.TEMPLATE_PATH, keep_vba=True)


class TestEscreverLinha:

    def test_colunas_literais_e_formulas(self, template):
        ws = template['RegimeEspecial']
        p._escrever_linha_regime_especial(ws, 5, _item_base())

        assert ws['S5'].value == 1.0                      # frete
        assert ws['T5'].value == 2.5                      # seguro + desp. acess.
        assert ws['U5'].value == 3.3                      # IPI
        assert ws['V5'].value == 0.5                      # desconto
        assert ws['Z5'].value == 0.3824                   # MVA fracionário intacto
        assert ws['AA5'].value == 4                       # MOD como int
        assert ws['AD5'].value.startswith('=IF(AA5=0')    # fórmula viva BC apurado
        assert ws['AP5'].value == '=IF(W5=0,0,X5/W5)'     # alíquota do produto
        assert ws['AP5'].number_format == '0.00%'
        assert ws['A5'].value == 2026 and ws['B5'].value == 5  # ano/mês da emissão

    def test_mva_percentual_cheio_e_normalizado_na_coluna_z(self, template):
        ws = template['RegimeEspecial']
        p._escrever_linha_regime_especial(ws, 6, _item_base(mva=38.24))
        assert ws['Z6'].value == 0.3824, (
            'MVA digitado como percentual cheio deve ser corrigido antes de '
            'gravar na coluna Z (fórmula AD exige Z<=5)'
        )

    def test_mva_none_permanece_vazio(self, template):
        ws = template['RegimeEspecial']
        p._escrever_linha_regime_especial(ws, 7, _item_base(mva=None))
        assert ws['Z7'].value is None


class TestCabecalhosTemplate:
    """Armadilha conhecida: a Excel Table guarda os nomes de coluna em metadata
    própria (tableColumns[i].name), separada do texto da célula do cabeçalho.
    Se divergirem, o Excel acusa 'conteúdo ilegível'."""

    def test_celulas_e_metadata_da_tabela_sincronizados(self, template):
        ws = template['RegimeEspecial']
        tabela = ws.tables['tabRegEsp']
        for i, col in enumerate(tabela.tableColumns, start=1):
            celula = ws.cell(row=4, column=i).value
            assert celula == col.name, (
                f'Coluna {i}: célula do cabeçalho ({celula!r}) difere do metadata '
                f'da tabela ({col.name!r}) — atualizar os DOIS ao renomear (ver AGENTS.md)'
            )

    def test_colunas_renomeadas_e_nova(self, template):
        ws = template['RegimeEspecial']
        assert ws['T4'].value == 'SEGURO/DESP. ACESS.'
        assert ws['U4'].value == 'IPI'
        assert ws['AP4'].value == 'ALÍQUOTA DO PRODUTO'


class TestGerarPlanilhaPontaAPonta:

    @pytest.fixture
    def conn_memoria(self):
        conn = sqlite3.connect(':memory:')
        conn.execute('''
            CREATE TABLE nfe_item_apuracao (
                chave_acesso TEXT, num_item INTEGER, mes_ano TEXT,
                tipo_planilha TEXT DEFAULT 'Original',
                num_nf TEXT, data_emissao TEXT, cnpj_remetente TEXT,
                razao_remetente TEXT, uf_remetente TEXT, cest TEXT, ncm TEXT,
                ean TEXT, cfop TEXT, cod_produto_origem TEXT,
                descricao_produto TEXT, unidade TEXT, quantidade REAL,
                valor_unitario REAL, frete REAL, seguro REAL, ipi_despesas REAL,
                desconto REAL, bc_icms_origem REAL, vlr_icms_origem REAL,
                pmc REAL, pmc_cmed REAL, mva REAL, mod_bc_icms_st TEXT,
                bc_icms_st_nfe REAL, vlr_icms_st_nfe REAL,
                PRIMARY KEY (chave_acesso, num_item, tipo_planilha)
            )
        ''')
        yield conn
        conn.close()

    def _inserir(self, conn, item, mes_ano='2026-05'):
        conn.execute(
            '''INSERT INTO nfe_item_apuracao
               (chave_acesso, num_item, mes_ano, tipo_planilha, num_nf,
                data_emissao, cnpj_remetente, razao_remetente, uf_remetente,
                cest, ncm, ean, cfop, cod_produto_origem, descricao_produto,
                unidade, quantidade, valor_unitario, frete, seguro,
                ipi_despesas, desconto, bc_icms_origem, vlr_icms_origem,
                pmc, pmc_cmed, mva, mod_bc_icms_st, bc_icms_st_nfe, vlr_icms_st_nfe)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (item['chave_acesso'], item['num_item'], mes_ano, 'Original',
             item['num_nf'], item['data_emissao'], item['cnpj_remetente'],
             item['razao_remetente'], item['uf_remetente'], item['cest'],
             item['ncm'], item['ean'], item['cfop'], item['cod_produto_origem'],
             item['descricao_produto'], item['unidade'], item['quantidade'],
             item['valor_unitario'], item['frete'], item['seguro'],
             item['ipi_despesas'], item['desconto'], item['bc_icms_origem'],
             item['vlr_icms_origem'], item['pmc'], item['pmc_cmed'],
             item['mva'], item['mod_bc_icms_st'], item['bc_icms_st_nfe'],
             item['vlr_icms_st_nfe']),
        )
        conn.commit()

    def test_gerar_planilha_dois_itens(self, conn_memoria, tmp_path, monkeypatch):
        self._inserir(conn_memoria, _item_base())
        self._inserir(conn_memoria, _item_base(num_item=2, mva=60.0))  # percentual cheio
        monkeypatch.setattr(p, 'ROOT', str(tmp_path))  # saída vai para pasta temporária

        caminho = p.gerar_planilha(conn_memoria, '2026-05', tipo_planilha='Original')

        assert os.path.exists(caminho)
        assert os.path.basename(caminho) == 'RegimeEspecialMS_2026_05_Original.xlsm'

        wb = openpyxl.load_workbook(caminho, keep_vba=True)
        ws = wb['RegimeEspecial']
        assert ws.tables['tabRegEsp'].ref == 'A4:AP6'   # 2 linhas de dados, até coluna AP
        assert ws['Z5'].value == 0.3824
        assert ws['Z6'].value == 0.6                     # 60 normalizado
        assert ws['AP6'].value == '=IF(W6=0,0,X6/W6)'
        assert wb['Resumo']['B6'].value.year == 2026 and wb['Resumo']['B6'].value.month == 5

    def test_competencia_vazia_gera_planilha_sem_dados(self, conn_memoria, tmp_path, monkeypatch):
        monkeypatch.setattr(p, 'ROOT', str(tmp_path))
        caminho = p.gerar_planilha(conn_memoria, '2026-06', tipo_planilha='Original')
        wb = openpyxl.load_workbook(caminho, keep_vba=True)
        assert wb['RegimeEspecial'].tables['tabRegEsp'].ref == 'A4:AP5'  # tabela mínima
